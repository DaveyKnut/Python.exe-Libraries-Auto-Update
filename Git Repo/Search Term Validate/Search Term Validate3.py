import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import re
import os
import sys
import subprocess
import re
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill
from tkinter import scrolledtext

class EnhancedSearchTermValidator:
    def __init__(self, master):
        self.master = master
        master.title("Search Term Validate")
        
        self.create_widgets()
        self.bind_events()
        
        self.stop_words = set([
            'a', 'about', 'after', 'all', 'also', 'an', 'and', 'another', 'any', 'are', 'as', 'at',
            'be', 'because', 'been', 'before', 'being', 'between', 'both', 'but', 'by',
            'came', 'can', 'come', 'could',
            'did', 'do',
            'each', 'even',
            'for', 'from', 'further', 'furthermore',
            'get', 'got',
            'had', 'has', 'have', 'he', 'her', 'here', 'hi', 'him', 'himself', 'his', 'how', 'however',
            'i', 'if', 'in', 'indeed', 'into', 'is', 'it', 'its',
            'just',
            'like',
            'made', 'many', 'me', 'might', 'more', 'moreover', 'most', 'much', 'must', 'my',
            'never', 'not', 'now',
            'of', 'on', 'only', 'or', 'other', 'our', 'out', 'over',
            'said', 'same', 'see', 'she', 'should', 'since', 'some', 'still', 'such',
            'take', 'than', 'that', 'the', 'their', 'them', 'then', 'there', 'therefore', 'these',
            'they', 'this', 'those', 'through', 'thus', 'to', 'too',
            'under', 'up',
            'very',
            'was', 'way', 'we', 'well', 'were', 'what', 'when', 'where', 'which', 'while', 'who',
            'will', 'with', 'would',
            'you', 'your'
        ])
        
        self.add_copyright_notice()
        
    def create_widgets(self):
        self.main_frame = ttk.Frame(self.master)
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        frame_input = ttk.Frame(self.main_frame, padding="10")
        frame_input.pack(fill=tk.BOTH, expand=True)
        
        self.text_input = tk.Text(frame_input, width=80, height=10, wrap=tk.WORD)
        self.text_input.pack(fill=tk.BOTH, expand=True)
        
        self.btn_check = ttk.Button(frame_input, text="Check", command=self.check_terms)
        self.btn_check.pack(pady=5)
        
        frame_stats = ttk.Frame(self.main_frame, padding="10")
        frame_stats.pack(fill=tk.BOTH, expand=True)
        
        self.stats_text = tk.Text(frame_stats, width=80, height=8, wrap=tk.WORD, state='disabled')
        self.stats_text.pack(fill=tk.BOTH, expand=True)
        
        frame_output = ttk.Frame(self.main_frame, padding="10")
        frame_output.pack(fill=tk.BOTH, expand=True)
        
        self.output_text = tk.Text(frame_output, width=80, height=10, wrap=tk.WORD, state='disabled')
        self.output_text.pack(fill=tk.BOTH, expand=True)
        
        frame_suggestions = ttk.Frame(self.main_frame, padding="10")
        frame_suggestions.pack(fill=tk.BOTH, expand=True)
        
        self.suggestions_text = tk.Text(frame_suggestions, width=80, height=10, wrap=tk.WORD, state='disabled')
        self.suggestions_text.pack(fill=tk.BOTH, expand=True)
        
        self.btn_export = ttk.Button(frame_output, text="Export to Excel", command=self.export_to_excel)
        self.btn_export.pack(pady=5)
        
        self.btn_open_report = ttk.Button(frame_output, text="Open Report", command=self.open_report, state='disabled')
        self.btn_open_report.pack(pady=5)
        
        self.regex_mode = tk.BooleanVar()
        self.chk_regex = ttk.Checkbutton(frame_input, text="Regex Mode", variable=self.regex_mode)
        self.chk_regex.pack(pady=5)
        
        # Add radio button for suffix wildcard suggestions
        self.suggest_wildcards = tk.BooleanVar()
        self.radio_wildcards = ttk.Radiobutton(frame_input, text="Suggest Suffix Wildcards", 
                                               variable=self.suggest_wildcards, value=True)
        self.radio_no_wildcards = ttk.Radiobutton(frame_input, text="No Suffix Wildcard Suggestions", 
                                                  variable=self.suggest_wildcards, value=False)
        self.radio_wildcards.pack(pady=5)
        self.radio_no_wildcards.pack(pady=5)
        
        # Add README button
        self.btn_readme = ttk.Button(frame_input, text="README", command=self.open_readme)
        self.btn_readme.pack(side=tk.RIGHT, padx=5, pady=5)
        
    def bind_events(self):
        self.text_input.bind('<KeyRelease>', self.on_key_release)
        
    def on_key_release(self, event):
        self.apply_syntax_highlighting()
        self.real_time_validation()
        
    def validate_search_terms(self, terms):
        results = []
        for term in terms.splitlines():
            if self.regex_mode.get():
                result = self.validate_regex(term)
            else:
                result = self.validate_boolean_search(term)
            results.append(result)
        return results
    
    def validate_regex(self, term):
        try:
            re.compile(term)
            return (term, "Valid regex", None, None)
        except re.error as e:
            suggestion = self.suggest_regex_fix(term, str(e))
            return (term, f"Invalid regex: {str(e)}", suggestion, None) 
        
    def suggest_regex_fix(self, term, error):
        if "unbalanced parenthesis" in error.lower():
            return f"Check your parentheses in: {term}"
        elif "unterminated character set" in error.lower():
            return f"You might be missing a closing bracket ']' in: {term}"
        return "Please check your regex syntax"
        
    def validate_boolean_search(self, term):
        errors = []
        warnings = []
        
        if len(term) > 455:
            errors.append("Too long (> 455 characters)")
        
        if term.count('(') != term.count(')'):
            errors.append("Unbalanced parentheses")
        
        if term.count('"') % 2 != 0:
            errors.append("Unmatched quotes")
        
        if re.search(r'[^\x00-\x7F]', term):
            warnings.append("Foreign character may not be searched for, check DT index")
        
        special_chars = ['&', '?', '#', '@', '$', '€', '£', '¥', '.']
        for char in special_chars:
            if char in term:
                warnings.append(f"Special character '{char}' may not be searchable, check DT index")
        
        if re.search(r'\b(and|or)\b', term):
            warnings.append("Lowercase 'and'/'or' needs speech marks around it to be searchable. Check DT index")
        
        stop_words_found = [word for word in term.lower().split() if word in self.stop_words]
        if stop_words_found:
            warnings.append(f"Stop word(s) detected: {', '.join(stop_words_found)}. Check DT Index to see if searchable")
        
        if errors:
            status = "Error: " + "; ".join(errors)
        elif warnings:
            status = "Warning: " + "; ".join(warnings)
        else:
            status = "Valid"
        
        suggestion, edited_term = self.suggest_fix(term, errors + warnings)
        
        # Add suffix wildcard suggestion if enabled
        if self.suggest_wildcards.get():
            wildcard_suggestion = self.suggest_suffix_wildcard(term)
            if wildcard_suggestion:
                if suggestion:
                    suggestion += f"; Suggested wildcard(s): {wildcard_suggestion}"
                else:
                    suggestion = f"Suggested wildcard(s): {wildcard_suggestion}"
                if edited_term is None:
                    edited_term = term  # Keep the original term, just suggest wildcards
        
        return (term, status, suggestion, edited_term)

    def suggest_suffix_wildcard(self, term):
        words = re.findall(r'\b[a-zA-Z]+\b', term)  # Only consider alphabetic words
        suggestions = []

        for word in words:
            # Skip short words and common words that don't benefit from wildcards
            if len(word) < 5 or word.lower() in self.stop_words:
                continue
            
            # Check for words already ending with a wildcard
            if word.endswith('*'):
                continue

            # Suggest wildcards for words that might have multiple endings
            if word.endswith(('s', 'ed', 'ing')):
                # Remove the ending and add wildcard
                base = re.sub(r'(s|ed|ing)$', '', word)
                if len(base) >= 4:  # Ensure the base is still long enough
                    suggestions.append(f"{base}*")
            elif len(word) >= 5:
                # For longer words, suggest a wildcard to catch potential variations
                suggestions.append(f"{word}*")

        if suggestions:
            return " ".join(suggestions)
        return None
            
    def suggest_fix(self, term, issues):
        suggestions = []
        edited_term = term
        
        if "Unbalanced parentheses" in issues:
            open_count = term.count('(')
            close_count = term.count(')')
            if open_count > close_count:
                diff = open_count - close_count
                suggestions.append(f"Add {diff} closing parenthesis/es")
                edited_term += ')' * diff
            else:
                diff = close_count - open_count
                suggestions.append(f"Add {diff} opening parenthesis/es")
                edited_term = '(' * diff + edited_term
        
        if "Unmatched quotes" in issues:
            suggestions.append("Add a closing quote")
            edited_term += '"'
        
        if "Too long (> 455 characters)" in issues:
            words = term.split()
            mid = len(words) // 2
            suggestions.append(f"Split into two terms")
            edited_term = f"1. {' '.join(words[:mid])}\n2. {' '.join(words[mid:])}"
        
        if "Lowercase 'and'/'or' needs speech marks" in ' '.join(issues):
            edited_term = re.sub(r'\b(and|or)\b', lambda m: f'"{m.group(0)}"', edited_term, flags=re.IGNORECASE)
            suggestions.append("Added quotes around lowercase 'and'/'or'")
        
        return ("; ".join(suggestions) if suggestions else None, edited_term if edited_term != term else None)
        
        
    def apply_syntax_highlighting(self):
        self.text_input.tag_remove('valid', '1.0', tk.END)
        self.text_input.tag_remove('invalid', '1.0', tk.END)
        self.text_input.tag_remove('warning', '1.0', tk.END)
        self.text_input.tag_remove('operator', '1.0', tk.END)
        self.text_input.tag_remove('parenthesis', '1.0', tk.END)
        self.text_input.tag_remove('special_char', '1.0', tk.END)
        self.text_input.tag_remove('stop_word', '1.0', tk.END)
        
        terms = self.text_input.get("1.0", tk.END).splitlines()
        for i, term in enumerate(terms, start=1):
            result = self.validate_boolean_search(term)
            
            if result[1].startswith("Error"):
                self.text_input.tag_add('invalid', f"{i}.0", f"{i}.end")
            elif result[1].startswith("Warning"):
                self.text_input.tag_add('warning', f"{i}.0", f"{i}.end")
            else:
                self.text_input.tag_add('valid', f"{i}.0", f"{i}.end")
            
            for match in re.finditer(r'\b(AND|OR|NOT)\b', term, re.IGNORECASE):
                start, end = match.span()
                self.text_input.tag_add('operator', f"{i}.{start}", f"{i}.{end}")
            
            for match in re.finditer(r'[()]', term):
                start, end = match.span()
                self.text_input.tag_add('parenthesis', f"{i}.{start}", f"{i}.{end}")
            
            for match in re.finditer(r'[*"''""„‚""''‛‛‚''""•√πƒ∂ß∆˚¬…¬˙©ƒ∂ß∆˙∫√ç≈ΩµΩ≈ç√∫˜µ≤≥≠&?#@$€£¥.]', term):
                start, end = match.span()
                self.text_input.tag_add('special_char', f"{i}.{start}", f"{i}.{end}")
            
            for word in term.lower().split():
                if word in self.stop_words:
                    start = term.lower().index(word)
                    end = start + len(word)
                    self.text_input.tag_add('stop_word', f"{i}.{start}", f"{i}.{end}")

        self.text_input.tag_config('valid', foreground='green')
        self.text_input.tag_config('invalid', foreground='red')
        self.text_input.tag_config('warning', foreground='orange')
        self.text_input.tag_config('operator', foreground='blue', font=('TkDefaultFont', 10, 'bold'))
        self.text_input.tag_config('parenthesis', foreground='purple', font=('TkDefaultFont', 10, 'bold'))
        self.text_input.tag_config('special_char', foreground='red', font=('TkDefaultFont', 10, 'bold'))
        self.text_input.tag_config('stop_word', background='yellow')
        
    def real_time_validation(self):
        terms = self.text_input.get("1.0", tk.END).strip()
        results = self.validate_search_terms(terms)
        self.display_results(results)
        self.update_stats(results)
        
    def check_terms(self):
        terms = self.text_input.get("1.0", tk.END).strip()
        if not terms:
            messagebox.showerror("Input Error", "Please input search terms")
            return
        results = self.validate_search_terms(terms)
        self.display_results(results)
        self.apply_syntax_highlighting()
        self.update_stats(results)

    def update_stats(self, results):
        total_terms = len(results)
        unique_terms = len(set(term for term, _, _, _ in results))
        duplicate_terms = total_terms - unique_terms
        
        error_terms = sum(1 for _, status, _, _ in results if status.startswith("Error"))
        valid_terms = unique_terms - error_terms
        warning_terms = sum(1 for _, status, _, _ in results if status.startswith("Warning"))
        
        warning_error_categories = Counter()
        for _, status, _, _ in results:
            if status.startswith("Warning") or status.startswith("Error"):
                categories = status.split(": ")[1].split("; ")
                warning_error_categories.update(categories)
        
        stats_text = f"""
Total terms: {total_terms}
Unique terms: {unique_terms}
Duplicate terms: {duplicate_terms}
Valid terms: {valid_terms}
Terms with warnings: {warning_terms}
Terms with errors: {error_terms}

Warning and Error Categories:
"""
        for category, count in warning_error_categories.most_common():
            stats_text += f"{category}: {count}\n"
        
        self.stats_text.config(state='normal')
        self.stats_text.delete("1.0", tk.END)
        self.stats_text.insert(tk.END, stats_text)
        self.stats_text.config(state='disabled')

    def display_results(self, results):
        self.output_text.config(state='normal')
        self.output_text.delete("1.0", tk.END)
        self.suggestions_text.config(state='normal')
        self.suggestions_text.delete("1.0", tk.END)
        
        for term, status, suggestion, edited_term in results:
            self.output_text.insert(tk.END, f"Term: {term}\nStatus: {status}\n\n")
            if suggestion or edited_term:
                self.suggestions_text.insert(tk.END, f"Term: {term}\n")
                if suggestion:
                    self.suggestions_text.insert(tk.END, f"Suggestion: {suggestion}\n")
                if edited_term:
                    self.suggestions_text.insert(tk.END, f"Edited Term: {edited_term}\n")
                self.suggestions_text.insert(tk.END, "\n")
        
        self.output_text.config(state='disabled')
        self.suggestions_text.config(state='disabled')
        
    def export_to_excel(self):
        terms = self.text_input.get("1.0", tk.END).strip()
        results = self.validate_search_terms(terms)
        
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if filepath:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Search Term Validation"
            
            headers = ["Search Term", "Status", "Suggestion", "Edited Term"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            for row, (term, status, suggestion, edited_term) in enumerate(results, start=2):
                ws.cell(row=row, column=1, value=term)
                ws.cell(row=row, column=2, value=status)
                ws.cell(row=row, column=3, value=suggestion if suggestion else "")
                ws.cell(row=row, column=4, value=edited_term if edited_term else "")
                
                if status == "Valid":
                    ws.cell(row=row, column=2).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif "Warning" in status:
                    ws.cell(row=row, column=2).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    ws.cell(row=row, column=2).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Set specific widths for columns A, B, C, and D
            ws.column_dimensions['A'].width = 150
            ws.column_dimensions['B'].width = 100
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 100
            
            wb.save(filepath)
            messagebox.showinfo("Export", "Results exported successfully")
            self.btn_open_report.config(state='normal')
            self.last_export_path = filepath

    def open_report(self):
        if hasattr(self, 'last_export_path') and os.path.exists(self.last_export_path):
            if os.name == 'nt':  # For Windows
                os.startfile(self.last_export_path)
            elif os.name == 'posix':  # For macOS and Linux
                subprocess.call(('open', self.last_export_path))
        else:
            messagebox.showerror("Error", "No recent export found or file not accessible.")

    def add_copyright_notice(self):
        copyright_label = ttk.Label(self.master, text="© Lewis Bennett, 2024", font=("TkDefaultFont", 8))
        copyright_label.pack(side=tk.BOTTOM, anchor=tk.SW, padx=5, pady=5)

    def open_readme(self):
        readme_window = tk.Toplevel(self.master)
        readme_window.title("README - Search Term Validate")
        readme_window.geometry("800x600")

        readme_text = scrolledtext.ScrolledText(readme_window, wrap=tk.WORD, font=("TkDefaultFont", 10))
        readme_text.pack(expand=True, fill='both', padx=10, pady=10)

        readme_content = """© Lewis Bennett, 2024. All rights reserved.
# Search Term Validate

## Table of Contents

1. [Introduction](#introduction)
2. [Installation](#installation)
3. [Usage](#usage)
4. [Features](#features)
5. [Syntax Analysis](#syntax-analysis)
6. [Wildcard Suggestion Feature](#wildcard-suggestion-feature)
7. [Grammar Rules and Examples](#grammar-rules-and-examples)
8. [Exporting Results](#exporting-results)
9. [Copyright Notice](#copyright-notice)

## Introduction

Search Term Validate is a powerful tool designed to help users validate and optimize their search terms. It provides real-time feedback on the validity of search terms, offers suggestions for improvements, and includes features like syntax highlighting and wildcard suggestions.

## Installation

1. Run the .exe installer  OR
2. Ensure you have Python 3.x installed on your system.
3. Install the required libraries:
 
   pip install tkinter openpyxl

3. Download and run (open with: Python) the `Search_Term_Validate3.py` file.

## Usage

1. Run the .exe or script:

   python Search_Term_Validate3.py

2. The main window of the application will open.
3. Enter your search terms in the top text box, one per line.
4. The tool will provide real-time feedback as you type.
5. Click the "Check" button to perform a full validation of all terms.
6. View results in the output boxes below. The three boxes are: 1: Enter terms, 2: Validates terms, 3: Provides suggestions

## Features

1. **Real-time Validation**: The tool validates search terms as you type, providing immediate feedback.

2. **Syntax Highlighting**: Different elements of the search terms are highlighted in various colors for easy identification. Red for error terms, orange for warnings, green for valid.

3. **Error and Warning Detection**: The tool identifies potential issues with search terms and categorizes them as errors or warnings. An error term will not run in Relativity. A warning term will run but you should consider the warning.

4. **Suggestions**: For problematic terms, the tool offers suggestions for improvement. !!These should not be blindly accepted, QC the suggestions and edit them as required!!

5. **Regex Mode**: Toggle between boolean search and regex validation.

6. **Wildcard Suggestions**: Option to suggest suffix wildcards for applicable terms. (Wildcards are suggested for nouns and verbs where a wildcard suffix would also return plurals and other word ending. For words ending in 's', 'ed', or 'ing', it suggests removing these endings and adding a wildcard.

7. **Statistics**: Provides an overview of the validation results, including counts of valid, warning, and error terms.

8. **Export to Excel**: Allows exporting of results to a formatted Excel file for further analysis.


## Syntax Analysis

The tool analyzes search terms based on the following criteria:

1. **Length**: Terms longer than 455 characters are flagged as errors. (These would error in Relativity)

2. **Parentheses Balance**: Checks for matching opening and closing parentheses. (These would error in Relativity) 

3. **Quotation Marks**: Ensures quotation marks are properly paired. (These would error in Relativity)

4. **Special Characters**: Identifies potentially unsearchable special characters.(These are the special characters in the DT Index which are not searchable by default)

5. **Boolean Operators**: Highlights AND, OR, NOT operators (to help readability of the terms).

6. **Stop Words**: Identifies DT Search default stop words that are not searchable by default without " ".

7. **Foreign Characters**: Warns about non-ASCII characters that might not be indexed (DT Index is foreign character "blind", so if you want to specifically search a foreign character, you will need to change the DT Index settings.


## Wildcard Suggestion Feature

The wildcard suggestion feature aims to optimize search terms by suggesting suffix wildcards. Here's how it works:

1. It analyzes words in the search term that are 5 characters or longer.

2. It ignores common stop words and words already ending with a wildcard.

3. For words ending in 's', 'ed', or 'ing', it suggests removing these endings and adding a wildcard.

4. For other long words, it suggests adding a wildcard to catch potential variations.


Example:
- Input: "Car causes crash"

- Suggestion: "Car* cause* crash*"

This feature helps to broaden searches and catch variations of words, potentially improving search results.

## Grammar Rules and Examples

1. **Phrase Searching**:

   - The tool identifies stop / noise words from the Relativity index that are not in quotation marks.

2. **Parentheses**:

   - The tool checks the proper use of parentheses (brackets) and " "

3. **Wildcards**:

   - The tool suggests the use of * at the end of a word to match any ending.

   - Example: run* (matches run, running, runner, etc.)

4. **Special Characters**:

   - The tool identifies terms containing special characters that are not indexed in Relativity &, #, @, etc., 

5. **Stop Words**:

   - Stop words are identified within terms 'a', 'the', 'in', etc as they may not be indexed in Relativity.

6. **Case Sensitivity**:

   - The tool treats search terms as case-insensitive.

## Exporting Results

1. Click the "Export to Excel" button after validating your terms.

2. Choose a location and filename for your Excel file.

3. The exported file will contain:

   - Original search terms
   - Validation status
   - Suggestions (if any)
   - Edited terms (if applicable)
   
4. The Excel file uses color coding:

   - Green: Valid terms
   - Yellow: Terms with warnings
   - Red: Terms with errors

## Copyright Notice

© Lewis Bennett, 2024. All rights reserved.

This software is provided for educational and research purposes only. Any unauthorized copying, modification, or distribution of this software is strictly prohibited.
"""

        readme_text.insert(tk.END, readme_content)
        readme_text.config(state='disabled')  # Make the text read-only

root = tk.Tk()
app = EnhancedSearchTermValidator(root)
root.mainloop()